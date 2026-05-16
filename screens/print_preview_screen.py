"""
print_preview_screen.py — 导出预览页面

功能:
- 顶部 TopBar 返回 record_list
- 选项栏：显示价格 CheckBox、显示人员 CheckBox
- 表格预览：表头 [序号, 日期, 工程名称, 施工单位, 机型, 价格, 上午, 下午, 加班, 合计时间, 合计金额, 备注, 开票人]
- 数据行交替背景色
- 统计行（合计）
- 按钮: '📊 导出Excel'、'📄 导出Word'、'📝 导出文本'
- 导出实现:
  - Excel: 使用 openpyxl，没有则 fallback 到 CSV
  - Word: 使用 zipfile 直接构建 docx XML
  - 导出到 /storage/emulated/0/Download/（Android）或~/Documents（桌面）
  - 文件名: 机械施工汇总_客户名_YYYYMMDD_HHMM.xlsx
"""

from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox
from kivy.uix.widget import Widget
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.metrics import dp
from kivy.clock import Clock
from datetime import datetime
import os
import csv

from config import get_theme, Size, AppInfo
from database import get_db
from widgets import Card, TopBar, BtnPrimary, BtnDanger, BtnOutline, ScrollList, Divider, draw_bg


# ──────────────────────────────────────────
# 表头定义
# ──────────────────────────────────────────
TABLE_HEADERS = [
    '序号', '日期', '工程名称', '施工单位', '机型', '价格',
    '上午', '下午', '加班', '合计时间', '合计金额', '备注', '开票人'
]


class PrintPreviewScreen(Screen):
    """导出预览页面"""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'print_preview'

        # 数据
        self._records = []
        self._customer_name = ''

        self._build_ui()

    # ──────────────────────────────────────────
    # UI 构建
    # ──────────────────────────────────────────

    def _build_ui(self):
        self._theme = get_theme()
        main = BoxLayout(orientation='vertical', spacing=0)
        self.add_widget(main)

        # 顶部导航
        top = TopBar(
            title='📄 导出文档',
            on_back=lambda w: self._go_back()
        )
        main.add_widget(top)

        # 表格标题栏
        hint = BoxLayout(
            orientation='horizontal',
            size_hint_y=None,
            height=dp(40),
            padding=(dp(Size.PAGE_PAD), dp(6)),
        )
        Clock.schedule_once(lambda dt: draw_bg(hint, get_theme().CARD), 0)
        hint_lbl = Label(
            text='📋 数据预览',
            font_size=dp(Size.BODY),
            bold=True,
            halign='left',
            size_hint_x=1,
        )
        Clock.schedule_once(lambda dt: setattr(hint_lbl, 'color', get_theme().TEXT), 0)
        hint.add_widget(hint_lbl)
        main.add_widget(hint)

        # 表格区域（可垂直+水平滚动）
        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=True, do_scroll_y=True)
        self._table_container = BoxLayout(
            orientation='vertical',
            size_hint_y=None,
            size_hint_x=None,
            spacing=0,
        )
        self._table_container.bind(minimum_height=self._table_container.setter('height'))
        self._table_container.bind(minimum_width=self._table_container.setter('width'))
        scroll.add_widget(self._table_container)
        main.add_widget(scroll)

        # 底部按钮栏
        btn_bar = BoxLayout(
            orientation='horizontal',
            size_hint_y=None,
            height=dp(56),
            padding=(dp(Size.PAGE_PAD), dp(8), dp(Size.PAGE_PAD), dp(8)),
            spacing=dp(8),
        )
        Clock.schedule_once(lambda dt: draw_bg(btn_bar, get_theme().TOOLBAR_BG), 0)

        export_excel = BtnPrimary(text='📊 导出Excel', color_name='GREEN')
        export_excel.bind(on_release=lambda w: self.do_export('xlsx'))
        btn_bar.add_widget(export_excel)

        export_word = BtnPrimary(text='📄 导出Word', color_name='GREEN')
        export_word.bind(on_release=lambda w: self.do_export('docx'))
        btn_bar.add_widget(export_word)

        export_text = BtnPrimary(text='📝 导出单据', color_name='GREEN')
        export_text.bind(on_release=lambda w: self._show_pdf_settings())
        btn_bar.add_widget(export_text)

        main.add_widget(btn_bar)

    # ──────────────────────────────────────────
    # PDF 单据导出
    # ──────────────────────────────────────────

    def _show_pdf_settings(self):
        """弹出 PDF 页面大小选择弹窗"""
        from kivy.uix.popup import Popup
        from kivy.uix.spinner import Spinner as Sp
        from widgets import BtnPrimary, BtnDanger
        theme = get_theme()
        content = BoxLayout(
            orientation='vertical',
            padding=dp(16),
            spacing=dp(12),
            size_hint=(None, None),
            size=(dp(280), dp(220)),
        )
        title_lbl = Label(
            text='导出PDF结算单',
            font_size=dp(Size.HEADING),
            bold=True,
            size_hint_y=None,
            height=dp(30),
            halign='center',
            color=theme.TEXT,
        )
        content.add_widget(title_lbl)

        content.add_widget(Label(
            text='选择纸张大小：',
            font_size=dp(Size.SMALL),
            size_hint_y=None,
            height=dp(22),
            color=theme.TEXT_SEC,
            halign='left',
        ))

        page_sizes = ['A4 (210x297mm)', 'A5 (148x210mm)', 'A6 (105x148mm)',
                      'B5 (176x250mm)', 'Letter (216x279mm)', '自定义']
        self._page_size_sp = Sp(
            text='A5 (148x210mm)',
            values=page_sizes,
            font_size=dp(Size.SMALL),
            size_hint_y=None,
            height=dp(40),
        )
        self._page_size_sp.background_color = theme.INPUT_BG
        content.add_widget(self._page_size_sp)

        btn_row = BoxLayout(
            orientation='horizontal',
            size_hint_y=None,
            height=dp(44),
            spacing=dp(10),
        )
        cancel_btn = BtnDanger(text='取消')
        cancel_btn.bind(on_release=lambda w: popup.dismiss())
        ok_btn = BtnPrimary(text='开始导出', color_name='GREEN')
        ok_btn.bind(on_release=lambda w: (
            popup.dismiss(),
            self._export_pdf(self._page_size_sp.text)
        ))
        btn_row.add_widget(cancel_btn)
        btn_row.add_widget(ok_btn)
        content.add_widget(btn_row)

        popup = Popup(
            title='',
            content=content,
            size_hint=(None, None),
            size=(dp(300), dp(260)),
            background='',
            background_color=(0, 0, 0, 0),
            separator_color=(0, 0, 0, 0),
        )
        from kivy.clock import Clock as _Clk
        _Clk.schedule_once(lambda dt: draw_bg(popup, theme.POPUP_BG, 12), 0)
        popup.open()

    def _export_pdf(self, page_size_str):
        """为每条记录生成一页独立PDF结算单"""
        if not self._records:
            self._show_toast('没有可导出的数据')
            return
        try:
            from fpdf import FPDF
            # 解析页面大小
            size_map = {
                'A4 (210x297mm)': ('A4', 'P'),
                'A5 (148x210mm)': ('A5', 'P'),
                'A6 (105x148mm)': ('A6', 'P'),
                'B5 (176x250mm)': ('B5', 'P'),
                'Letter (216x279mm)': ('Letter', 'P'),
                '自定义': ('A5', 'P'),
            }
            fmt, orient = size_map.get(page_size_str, ('A5', 'P'))

            dir_path = self._get_export_dir()
            ts = datetime.now().strftime('%Y%m%d_%H%M')
            name = self._customer_name or '未命名'
            safe_name = name.replace(' ', '_').replace('/', '_')
            pdf_path = os.path.join(dir_path, f'机械施工结算单_{safe_name}_{ts}.pdf')

            pdf = FPDF(orientation=orient, unit='mm', format=fmt)
            pdf.add_font('zh', '', 'C:/Windows/Fonts/msyh.ttc', uni=True)
            # 加粗字体，如果不存在则用普通字体代替
            try:
                pdf.add_font('zh', 'B', 'C:/Windows/Fonts/msyhbd.ttc', uni=True)
            except Exception:
                pdf.add_font('zh', 'B', 'C:/Windows/Fonts/msyh.ttc', uni=True)

            for record in self._records:
                self._add_pdf_page(pdf, record)
            pdf.output(pdf_path)
            self._show_toast(f'已导出: {pdf_path}')
        except Exception as e:
            self._show_toast(f'导出PDF失败: {str(e)}')

    def _add_pdf_page(self, pdf, record):
        """添加一页PDF结算单"""
        pdf.add_page()
        w = pdf.w - 20  # 可用宽度（减去边距）
        left = 10

        # 标题
        pdf.set_font('zh', 'B', 16)
        pdf.set_xy(left, 10)
        pdf.cell(w, 10, '机械施工结算单', border=0, align='C')

        # 分隔线
        pdf.set_draw_color(0, 100, 200)
        pdf.set_line_width(0.5)
        pdf.line(left, 22, left + w, 22)

        y = 28
        pdf.set_font('zh', '', 10)

        # 基本信息表格（两列）
        info_items = [
            ('客户名称', record.get('customer_name', '')),
            ('项目名称', record.get('project_name', '')),
            ('施工日期', record.get('date', '')),
            ('产品/机型', record.get('machine_type', '')),
            ('施工单位', record.get('construction_unit', '')),
        ]
        col_w = w / 2
        for i, (label, val) in enumerate(info_items):
            col = i % 2
            row_idx = i // 2
            cx = left + col * col_w
            cy = y + row_idx * 8
            pdf.set_xy(cx, cy)
            pdf.set_font('zh', '', 9)
            # 标签
            pdf.set_fill_color(230, 240, 255)
            pdf.cell(col_w * 0.3, 7, f'  {label}', border=1, fill=True)
            # 值
            pdf.set_fill_color(255, 255, 255)
            pdf.cell(col_w * 0.7, 7, f'  {val}', border=1, fill=True)

        y2 = y + (len(info_items) + 1) // 2 * 8 + 4

        # 时间明细表头
        pdf.set_xy(left, y2)
        pdf.set_fill_color(0, 100, 200)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('zh', 'B', 9)
        col_widths = [20, 20, 20, 20, 20, 30, 30]
        headers = ['上午开始', '上午结束', '下午开始', '下午结束', '加班开始', '加班结束', '合计']
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 7, h, border=1, fill=True, align='C')

        # 时间明细数据
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('zh', '', 9)
        pdf.set_xy(left, y2 + 7)
        am_s = record.get('am_start', '')
        am_e = record.get('am_end', '')
        pm_s = record.get('pm_start', '')
        pm_e = record.get('pm_end', '')
        ot_s = record.get('ot_start', '')
        ot_e = record.get('ot_end', '')
        total_h = record.get('total_hours', 0)
        pdf.set_fill_color(248, 248, 248)
        vals = [am_s, am_e, pm_s, pm_e, ot_s, ot_e, f'{total_h:.2f}h']
        for i, v in enumerate(vals):
            pdf.cell(col_widths[i], 7, v, border=1, align='C', fill=(i % 2 == 0))

        y3 = y2 + 14 + 4

        # 金额信息
        pdf.set_xy(left, y3)
        pdf.set_font('zh', 'B', 10)
        pdf.set_fill_color(255, 245, 220)
        price = record.get('price', 0)
        unit = record.get('unit', '')
        total_amt = record.get('total_amount', 0)
        pdf.cell(w * 0.4, 8, f'  单价: ¥{price:.2f}/{unit}', border=1, fill=True)
        pdf.set_fill_color(255, 235, 200)
        pdf.cell(w * 0.6, 8, f'  合计金额: ¥{total_amt:.2f}', border=1, fill=True, align='C')

        y4 = y3 + 10

        # 备注
        remarks = record.get('remarks', '')
        issuer = record.get('issuer', '')
        pdf.set_xy(left, y4)
        pdf.set_font('zh', '', 9)
        pdf.set_fill_color(245, 245, 245)
        pdf.cell(w * 0.6, 7, f'  备注: {remarks}', border=1, fill=True)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(w * 0.4, 7, f'  开票人: {issuer}', border=1, fill=True)

        y5 = y4 + 9

        # 底部信息
        pdf.set_xy(left, y5)
        pdf.set_font('zh', '', 8)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(w, 5, f'打印日期: {datetime.now().strftime("%Y-%m-%d %H:%M")}', align='R')

    # ──────────────────────────────────────────
    # 导航
    # ──────────────────────────────────────────

    def _go_back(self):
        """返回到记录列表页面"""
        if self.manager and self.manager.has_screen('record_list'):
            self.manager.current = 'record_list'

    # ──────────────────────────────────────────
    # 数据加载
    # ──────────────────────────────────────────

    def load_records(self, records, customer_name=''):
        """由外部调用，传入要导出的记录列表"""
        self._records = records
        self._customer_name = customer_name
        self.load_preview()

    def load_preview(self):
        """根据当前数据和选项刷新表格预览"""
        self._table_container.clear_widgets()

        if not self._records:
            empty_lbl = Label(
                text='暂无数据',
                font_size=dp(Size.BODY),
                size_hint_y=None,
                height=dp(80),
                halign='center',
            )
            Clock.schedule_once(lambda dt: setattr(empty_lbl, 'color', get_theme().TEXT_DIM), 0)
            self._table_container.add_widget(empty_lbl)
            return

        theme = get_theme()

        # 先构建所有行数据（二维列表）
        rows_data = self._build_rows_data()

        # 渲染表头
        headers = TABLE_HEADERS
        header_row = self._build_header_row(headers, theme)
        self._table_container.add_widget(header_row)

        # 分割线
        divider = Divider()
        self._table_container.add_widget(divider)

        # 渲染数据行（交替背景色）
        for idx, row_data in enumerate(rows_data):
            bg = theme.CARD if idx % 2 == 0 else theme.CARD_ALT
            row_widget = self._build_data_row(row_data, headers, bg, theme)
            self._table_container.add_widget(row_widget)

        # 统计行
        if rows_data:
            total_row = self._build_total_row(rows_data, headers, theme)
            self._table_container.add_widget(total_row)

    def _get_visible_headers(self):
        """始终显示所有列"""
        return TABLE_HEADERS

    def _build_rows_data(self):
        """从 self._records 构建二维数据列表"""
        rows = []
        for idx, r in enumerate(self._records, start=1):
            row = {
                '序号': str(idx),
                '日期': r.get('date', ''),
                '工程名称': r.get('project_name', ''),
                '施工单位': r.get('construction_unit', ''),
                '机型': r.get('machine_type', ''),
                '价格': self._format_price(r.get('price', 0), r.get('unit', '')),
                '上午': self._format_time_range(r.get('am_start', ''), r.get('am_end', '')),
                '下午': self._format_time_range(r.get('pm_start', ''), r.get('pm_end', '')),
                '加班': self._format_time_range(r.get('ot_start', ''), r.get('ot_end', '')),
                '合计时间': self._format_hours(r.get('total_hours', 0)),
                '合计金额': self._format_amount(r.get('total_amount', 0)),
                '备注': r.get('remarks', '') or '',
                '开票人': r.get('issuer', '') or '',
            }
            rows.append(row)
        return rows

    def _format_price(self, price, unit):
        """格式化单价"""
        try:
            return f"¥{float(price):.2f}/{unit}" if unit else f"¥{float(price):.2f}"
        except (ValueError, TypeError):
            return f"¥0.00"

    def _format_time_range(self, start, end):
        """格式化时间段"""
        s = (start or '').strip()
        e = (end or '').strip()
        if s and e:
            return f"{s}-{e}"
        elif s:
            return s
        return ''

    def _format_hours(self, hours):
        """格式化小时数"""
        try:
            h = float(hours)
            return f"{h:.1f}h"
        except (ValueError, TypeError):
            return '0.0h'

    def _format_amount(self, amount):
        """格式化金额"""
        try:
            return f"¥{float(amount):.2f}"
        except (ValueError, TypeError):
            return '¥0.00'

    # ──────────────────────────────────────────
    # 表格行构建
    # ──────────────────────────────────────────

    def _build_header_row(self, headers, theme):
        """构建表头行"""
        return self._make_table_row(headers, theme.ACCENT, (1, 1, 1, 1),
                                    bold=True, height=36)

    def _build_data_row(self, row_data, headers, bg, theme):
        """构建数据行"""
        cells = [row_data.get(h, '') for h in headers]
        return self._make_table_row(cells, bg, theme.TEXT, bold=False, height=28)

    def _build_total_row(self, rows_data, headers, theme):
        """构建统计合计行"""
        # 计算合计时间和合计金额
        total_hours = 0.0
        total_amount = 0.0
        for row in rows_data:
            # 解析合计时间
            h_str = row.get('合计时间', '0.0h')
            try:
                total_hours += float(h_str.replace('h', ''))
            except (ValueError, TypeError):
                pass
            # 解析合计金额
            a_str = row.get('合计金额', '¥0.00')
            try:
                total_amount += float(a_str.replace('¥', ''))
            except (ValueError, TypeError):
                pass

        cells = []
        for h in headers:
            if h == '序号':
                cells.append('')
            elif h == '合计时间':
                cells.append(self._format_hours(total_hours))
            elif h == '合计金额':
                cells.append(self._format_amount(total_amount))
            elif h in ('备注', '开票人', '上午', '下午', '加班', '价格'):
                cells.append('')
            else:
                cells.append('')

        return self._make_table_row(cells, theme.GOLD, (1, 1, 1, 1),
                                    bold=True, height=32)

    def _make_table_row(self, cells, bg_color, text_color, bold=False, height=28):
        """通用表格行构建，cells 为字符串列表，固定最小列宽实现可滚动"""
        theme = get_theme()
        row = BoxLayout(
            orientation='horizontal',
            size_hint_y=None,
            size_hint_x=None,
            width=dp(80 * len(cells)),  # 固定总宽度（每列80dp）
            height=dp(height),
            spacing=0,
        )
        draw_bg(row, bg_color)

        for cell_text in cells:
            lbl = Label(
                text=str(cell_text),
                font_size=dp(Size.SMALL if not bold else Size.BODY),
                bold=bold,
                color=text_color,
                size_hint_x=None,
                width=dp(80),  # 每列固定80dp
                halign='center',
                valign='middle',
                shorten=True,
                shorten_from='right',
            )
            row.add_widget(lbl)

        return row

    # ──────────────────────────────────────────
    # 导出入口
    # ──────────────────────────────────────────

    def do_export(self, fmt):
        """根据格式导出：'xlsx' / 'docx' / 'csv'"""
        if not self._records:
            self._show_toast('没有可导出的数据')
            return

        if fmt == 'xlsx':
            self._export_xlsx()
        elif fmt == 'docx':
            self._export_docx()
        elif fmt == 'csv':
            self._export_csv()

    # ──────────────────────────────────────────
    # 导出路径
    # ──────────────────────────────────────────

    def _get_export_dir(self):
        """根据平台获取导出目录（与备份目录相同）"""
        from config import AppInfo
        d = AppInfo.get_backup_dir()
        os.makedirs(d, exist_ok=True)
        return d

    def _get_export_filename(self, ext):
        """生成导出文件名: 机械施工汇总_客户名_YYYYMMDD_HHMM.ext"""
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        name = self._customer_name or '未命名'
        safe_name = name.replace(' ', '_').replace('/', '_')
        return f"机械施工汇总_{safe_name}_{ts}.{ext}"

    # ──────────────────────────────────────────
    # 导出 Excel (openpyxl) / fallback CSV
    # ──────────────────────────────────────────

    def _export_xlsx(self):
        """导出为 Excel (.xlsx)，没有 openpyxl 则 fallback 到 CSV"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            self._export_csv()
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '机械施工汇总'

            # 准备表头和数据
            headers = self._get_visible_headers()
            rows_data = self._build_rows_data()

            # 写入表头
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='385EA8', end_color='385EA8', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0'),
            )

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # 写入数据行
            data_font = Font(size=10)
            alt_fill = PatternFill(start_color='F2F6FC', end_color='F2F6FC', fill_type='solid')
            data_align = Alignment(horizontal='center', vertical='center')

            for row_idx, row_data in enumerate(rows_data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = row_data.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border
                    # 交替背景
                    if row_idx % 2 == 1:
                        cell.fill = alt_fill

            # 统计行
            total_row_idx = len(rows_data) + 2
            # 计算合计
            total_hours = sum(
                float(r.get('total_hours', 0) or 0) for r in self._records
            )
            total_amount = sum(
                float(r.get('total_amount', 0) or 0) for r in self._records
            )

            total_font = Font(bold=True, size=10, color='FFFFFF')
            total_fill = PatternFill(start_color='D4880F', end_color='D4880F', fill_type='solid')

            for col_idx, header in enumerate(headers, start=1):
                if header == '合计时间':
                    value = self._format_hours(total_hours)
                elif header == '合计金额':
                    value = self._format_amount(total_amount)
                else:
                    value = ''
                cell = ws.cell(row=total_row_idx, column=col_idx, value=value)
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = data_align
                cell.border = thin_border

            # 列宽自适应
            for col_idx, header in enumerate(headers, start=1):
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max(10, len(header) * 3)

            # 保存
            export_dir = self._get_export_dir()
            fname = self._get_export_filename('xlsx')
            fpath = os.path.join(export_dir, fname)
            wb.save(fpath)
            self._show_toast(f'✅ Excel 导出成功\n{fpath}')

        except Exception as e:
            self._show_toast(f'❌ Excel 导出失败: {e}')

    # ──────────────────────────────────────────
    # 导出 Word (docx via zipfile)
    # ──────────────────────────────────────────

    def _export_docx(self):
        """使用 zipfile 直接构建 docx XML 来导出 Word 文档"""
        try:
            import zipfile
            import xml.etree.ElementTree as ET
        except ImportError:
            self._show_toast('❌ 缺少 zipfile 模块')
            return

        try:
            headers = self._get_visible_headers()
            rows_data = self._build_rows_data()

            # 计算合计
            total_hours = sum(
                float(r.get('total_hours', 0) or 0) for r in self._records
            )
            total_amount = sum(
                float(r.get('total_amount', 0) or 0) for r in self._records
            )

            # 构建 Word 文档的 XML
            # docx 本质是一个 ZIP 包，包含多个 XML 文件
            # 这里用最简方式构建

            # 构建 document.xml
            body_parts = []

            # 标题段落
            title = f"机械施工汇总 — {self._customer_name or '未命名'}"
            body_parts.append(f'''
                <w:p>
                    <w:pPr>
                        <w:jc w:val="center"/>
                        <w:spacing w:before="200" w:after="200"/>
                    </w:pPr>
                    <w:r>
                        <w:rPr>
                            <w:b/>
                            <w:sz w:val="28"/>
                            <w:color w:val="385EA8"/>
                        </w:rPr>
                        <w:t>{self._xml_escape(title)}</w:t>
                    </w:r>
                </w:p>''')

            # 日期
            body_parts.append(f'''
                <w:p>
                    <w:pPr>
                        <w:jc w:val="right"/>
                        <w:spacing w:after="200"/>
                    </w:pPr>
                    <w:r>
                        <w:rPr>
                            <w:sz w:val="18"/>
                            <w:color w:val="808080"/>
                        </w:rPr>
                        <w:t>导出日期: {datetime.now().strftime('%Y-%m-%d %H:%M')}</w:t>
                    </w:r>
                </w:p>''')

            # 表格
            table_xml = '<w:tbl>'
            # 表格属性
            table_xml += '''
                <w:tblPr>
                    <w:tblStyle w:val="TableGrid"/>
                    <w:tblW w:w="5000" w:type="pct"/>
                    <w:jc w:val="center"/>
                </w:tblPr>'''

            # 表头行
            table_xml += '<w:tr>'
            for h in headers:
                table_xml += f'''
                    <w:tc>
                        <w:tcPr>
                            <w:shd w:val="clear" w:color="auto" w:fill="385EA8"/>
                            <w:tcW w:w="{int(5000 / len(headers))}" w:type="pct"/>
                        </w:tcPr>
                        <w:p>
                            <w:pPr>
                                <w:jc w:val="center"/>
                                <w:spacing w:before="40" w:after="40"/>
                            </w:pPr>
                            <w:r>
                                <w:rPr>
                                    <w:b/>
                                    <w:sz w:val="18"/>
                                    <w:color w:val="FFFFFF"/>
                                </w:rPr>
                                <w:t>{self._xml_escape(h)}</w:t>
                            </w:r>
                        </w:p>
                    </w:tc>'''
            table_xml += '</w:tr>'

            # 数据行
            for row_idx, row_data in enumerate(rows_data):
                bg_color = 'FFFFFF' if row_idx % 2 == 0 else 'F2F6FC'
                table_xml += '<w:tr>'
                for h in headers:
                    text = str(row_data.get(h, ''))
                    table_xml += f'''
                        <w:tc>
                            <w:tcPr>
                                <w:shd w:val="clear" w:color="auto" w:fill="{bg_color}"/>
                                <w:tcW w:w="{int(5000 / len(headers))}" w:type="pct"/>
                            </w:tcPr>
                            <w:p>
                                <w:pPr>
                                    <w:jc w:val="center"/>
                                    <w:spacing w:before="20" w:after="20"/>
                                </w:pPr>
                                <w:r>
                                    <w:rPr>
                                        <w:sz w:val="18"/>
                                    </w:rPr>
                                    <w:t>{self._xml_escape(text)}</w:t>
                                </w:r>
                            </w:p>
                        </w:tc>'''
                table_xml += '</w:tr>'

            # 合计行
            table_xml += '<w:tr>'
            for h in headers:
                if h == '合计时间':
                    text = self._format_hours(total_hours)
                elif h == '合计金额':
                    text = self._format_amount(total_amount)
                else:
                    text = ''
                table_xml += f'''
                    <w:tc>
                        <w:tcPr>
                            <w:shd w:val="clear" w:color="auto" w:fill="D4880F"/>
                            <w:tcW w:w="{int(5000 / len(headers))}" w:type="pct"/>
                        </w:tcPr>
                        <w:p>
                            <w:pPr>
                                <w:jc w:val="center"/>
                                <w:spacing w:before="40" w:after="40"/>
                            </w:pPr>
                            <w:r>
                                <w:rPr>
                                    <w:b/>
                                    <w:sz w:val="18"/>
                                    <w:color w:val="FFFFFF"/>
                                </w:rPr>
                                <w:t>{self._xml_escape(text)}</w:t>
                            </w:r>
                        </w:p>
                    </w:tc>'''
            table_xml += '</w:tr>'

            table_xml += '</w:tbl>'
            body_parts.append(table_xml)

            body_xml = '<w:body>' + '\n'.join(body_parts) + '</w:body>'

            document_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
            <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
                        xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                {body_xml}
            </w:document>'''

            # 构建 ZIP 包
            export_dir = self._get_export_dir()
            fname = self._get_export_filename('docx')
            fpath = os.path.join(export_dir, fname)

            with zipfile.ZipFile(fpath, 'w', zipfile.ZIP_DEFLATED) as zf:
                # [Content_Types].xml
                zf.writestr('[Content_Types].xml', '''<?xml version="1.0" encoding="UTF-8"?>
                <Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
                    <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
                    <Default Extension="xml" ContentType="application/xml"/>
                    <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
                </Types>''')

                # _rels/.rels
                zf.writestr('_rels/.rels', '''<?xml version="1.0" encoding="UTF-8"?>
                <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                    <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
                </Relationships>''')

                # word/_rels/document.xml.rels
                zf.writestr('word/_rels/document.xml.rels', '''<?xml version="1.0" encoding="UTF-8"?>
                <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                </Relationships>''')

                # word/document.xml
                zf.writestr('word/document.xml', document_xml)

                # word/styles.xml (minimal)
                zf.writestr('word/styles.xml', '''<?xml version="1.0" encoding="UTF-8"?>
                <w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
                    <w:style w:type="table" w:styleId="TableGrid">
                        <w:name w:val="Table Grid"/>
                        <w:pPr>
                            <w:spacing w:before="0" w:after="0"/>
                        </w:pPr>
                        <w:rPr>
                            <w:sz w:val="18"/>
                        </w:rPr>
                        <w:tblPr>
                            <w:tblBorders>
                                <w:top w:val="single" w:sz="4" w:space="0" w:color="D0D0D0"/>
                                <w:bottom w:val="single" w:sz="4" w:space="0" w:color="D0D0D0"/>
                                <w:left w:val="single" w:sz="4" w:space="0" w:color="D0D0D0"/>
                                <w:right w:val="single" w:sz="4" w:space="0" w:color="D0D0D0"/>
                            </w:tblBorders>
                        </w:tblPr>
                    </w:style>
                </w:styles>''')

            self._show_toast(f'✅ Word 导出成功\n{fpath}')

        except Exception as e:
            self._show_toast(f'❌ Word 导出失败: {e}')

    # ──────────────────────────────────────────
    # 导出 CSV (纯文本 fallback)
    # ──────────────────────────────────────────

    def _export_csv(self):
        """导出为 CSV 纯文本"""
        try:
            headers = self._get_visible_headers()
            rows_data = self._build_rows_data()

            export_dir = self._get_export_dir()
            fname = self._get_export_filename('csv')
            fpath = os.path.join(export_dir, fname)

            with open(fpath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)

                # 表头
                writer.writerow(headers)

                # 数据行
                for row_data in rows_data:
                    writer.writerow([row_data.get(h, '') for h in headers])

                # 合计行
                total_hours = sum(
                    float(r.get('total_hours', 0) or 0) for r in self._records
                )
                total_amount = sum(
                    float(r.get('total_amount', 0) or 0) for r in self._records
                )
                total_row = []
                for h in headers:
                    if h == '合计时间':
                        total_row.append(self._format_hours(total_hours))
                    elif h == '合计金额':
                        total_row.append(self._format_amount(total_amount))
                    else:
                        total_row.append('')
                writer.writerow(total_row)

            self._show_toast(f'✅ CSV 导出成功\n{fpath}')

        except Exception as e:
            self._show_toast(f'❌ CSV 导出失败: {e}')

    # ──────────────────────────────────────────
    # XML 转义工具
    # ──────────────────────────────────────────

    @staticmethod
    def _xml_escape(text):
        """转义 XML 特殊字符"""
        text = str(text)
        text = text.replace('&', '&amp;')
        text = text.replace('"', '&quot;')
        text = text.replace("'", '&apos;')
        text = text.replace('<', '&lt;')
        text = text.replace('>', '&gt;')
        return text

    # ──────────────────────────────────────────
    # Toast 提示
    # ──────────────────────────────────────────

    def _show_toast(self, msg):
        """简易 Toast 提示"""
        theme = get_theme()
        content = BoxLayout(padding=dp(20))
        lbl = Label(text=msg, font_size=dp(Size.SMALL),
                    halign='center')
        Clock.schedule_once(lambda dt: setattr(lbl, 'color', theme.TEXT), 0)
        content.add_widget(lbl)

        popup = Popup(
            title='',
            content=content,
            size_hint=(0.7, 0.25),
            background='',
            background_color=(0, 0, 0, 0),
            separator_color=(0, 0, 0, 0),
        )
        Clock.schedule_once(lambda dt: draw_bg(popup, theme.POPUP_BG, 8), 0)
        popup.open()
        Clock.schedule_once(lambda dt: popup.dismiss(), 2.0)
